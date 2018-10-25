import sys

from PyQt5.QtWidgets import *
from PyQt5.QtGui import *

# 读取原始数据，提取siteID，Lat，Long，以及所有的comments，存入到dict中
# 输出统一格式后的dict
def importExcel(workbookName = 'siteTemplate.xlsx',workbookSheet = 'Sheet1'):
		import openpyxl
		sites = {'MSL Released':{},'Cluster Finalization':{},'STAD Table Locked':{},'Equipment Install Complete':{},'RFI Report':{},'Commissioning and Integration':{}}
		subFolder = {'MSL Released':'MSL Released','Cluster Finalization':'Cluster Finalization','STAD Table Locked':'STAD Table Locked','Equipment Install Complete':'Equipment Install Complete','RFI Report':'RFI Report','Commissioning and Integration':'Commissioning and Integration'}
				
		#open Excel file
		wb = openpyxl.load_workbook(filename = workbookName)
		ws = wb[workbookSheet]
		
		#create a dictionary with all sites information
		for row in range(2,ws.max_row+1):
				if ws.cell(column = 27, row = row).value:				
						siteKey = str(ws.cell(column=1,row = row).value)
						siteID = str(ws.cell(column=1, row = row).value)
						siteDescription = ''
						for item in range(1,28):
								siteDescription += str(ws.cell(column = item, row =1).value) +' = ' + str(ws.cell(column = item, row = row).value) + '\n'
						siteCoordinates = str(ws.cell(column = 6, row = row).value) + ',' +  str(ws.cell(column=5, row = row).value) + ',0'
						sites[str(ws.cell(column = 27, row = row).value)][siteKey] = {'siteID': siteID, 'siteDescription': siteDescription, 'siteCoordinates': siteCoordinates}		
				
		return sites;


# 打开kml模板，将dict中的数据输入进去
# 使用xml.etree.ElementTree 作为操作xml(kml)的 lib
# 需要注意的是要先定义namespace
# 通过循环将dict中的数据，一层一层写入到xml(kml)中，最后储存
def generateSite (sites,kmlName):
	
		import xml.etree.ElementTree as ET


		#define default namespace in kml file
		ET.register_namespace('', "http://www.opengis.net/kml/2.2")
		ET.register_namespace('gx', "http://www.google.com/kml/ext/2.2")
		#ET.register_namespace('kml', "http://www.opengis.net/kml/2.2")
		ET.register_namespace('atom', "http://www.w3.org/2005/Atom")
		
		siteStyle = {'MSL Released':'#Site Style3','Equipment Install Complete':'#m_ylw-pushpin','Cluster Finalization':'#Site Style1','STAD Table Locked':'#Site Style6','RFI Report':'#Site Style7','Commissioning and Integration':'#Site Style10'}
		subFolder = {'MSL Released':'MSL Released','Cluster Finalization':'Cluster Finalization','STAD Table Locked':'STAD Table Locked','Equipment Install Complete':'Equipment Install Complete','RFI Report':'RFI Report','Commissioning and Integration':'Commissioning and Integration'}
		
		
		#open kml file
		tree = ET.parse ('NSW Small Cell Template03.kml')
		ns = {'nslink': "http://www.opengis.net/kml/2.2"}
		doc = tree.find('nslink:Document', ns)
		folder = doc.find('nslink:Folder',ns)
		
		#change file name as per input
		filename = doc.find('nslink:name',ns)
		if 	not (kmlName is None):
				filename.text = kmlName
		else:
				filename.text = 'newkml.kml'
		
		for item_id, item_value in sites.items():
				#create subfolder for 6 different categories
				subFolder[item_id] = ET.SubElement(folder,'Folder')
				name = ET.SubElement(subFolder[item_id],'name')
				name.text = str(item_id)
				
				for item_site_id, item_site_value in item_value.items():
										
						#create basic information for the pin
						placemark = ET.SubElement(subFolder[item_id],'Placemark')
						ID =ET.SubElement(placemark,'name')
						ID.text = sites[item_id][item_site_id]['siteID']
						description = ET.SubElement(placemark,'description')
						description.text = sites[item_id][item_site_id]['siteDescription']
						
						styleUrl = ET.SubElement(placemark,'styleUrl')
						styleUrl.text = siteStyle[item_id]
						
						point = ET.SubElement(placemark,'Point')
						coordinates = ET.SubElement(point,'coordinates')
						coordinates.text = sites[item_id][item_site_id]['siteCoordinates']
		
		#save kml file
		tree.write(kmlName, encoding = 'utf-8', xml_declaration=True)
		
		return; 
	

# 生成kml的主函数
def gKML():
		
		foo.form_widget.currentStatus= "Ready to start"
		#kmlName = 'NSW Small Cell_17082018test.kml'
		kmlName = 'NSW Small Cell_'+foo.form_widget.dateTypeIn.text()+'.kml'
		sites = importExcel()
		generateSite(sites,kmlName)
		foo.form_widget.currentStatus= "Completed"

# UI，用于提供按钮和原始文件名的输入
class MyMainWindow(QMainWindow):
		
		def __init__(self, parent = None):
				
				super(MyMainWindow, self).__init__(parent)
				self.form_widget = FormWidget(self)
				self.setCentralWidget(self.form_widget)
				
				self.abc = 'test01'
				
				#set Exit Button
				exitAct = QAction(QIcon('exit.png'), 'Exit', self)
				exitAct.setShortcut('Ctrl+Q')
				exitAct.setStatusTip('Exit application')
				exitAct.triggered.connect(self.close)
				
				#show StatusBar
				self.statusBar()
				
				#create MenuBar
				menubar = self.menuBar()
				fileMenu = menubar.addMenu('&File')
				toolDes = QAction('This is a tracker updating tool',self)
				fileMenu.addAction(toolDes)
				
				fileMenu.addAction(exitAct)
				
				#create ToolBar
				toolbar = self.addToolBar('Exit')
				toolbar.addAction(exitAct)
				
				#set popped up window location/title
				self.setGeometry(300,300,250,250)
				self.setWindowTitle('Daily Tracker Update')
		
class FormWidget(QWidget):
		
		def __init__(self,parent):
				
				super(FormWidget,self).__init__(parent)
				
				self.layout = QVBoxLayout(self)
				
				self.dateTypeIn = QLineEdit()
				self.layout.addWidget(self.dateTypeIn)
								
				self.button1 = QPushButton("Generate KML")
				self.layout.addWidget(self.button1)
				
				self.button1.clicked.connect(gKML)
				self.button1.clicked.connect(self.clickMethod)
																
				self.setLayout(self.layout)
				
				
		def clickMethod(self):
				QMessageBox.about(self,"Result","Done")

if __name__ == '__main__' :

		app = QApplication(sys.argv)
		foo = MyMainWindow()
		foo.show()
		sys.exit(app.exec_())
		
		
